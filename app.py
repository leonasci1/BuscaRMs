import streamlit as st
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.utils.cell import range_boundaries, get_column_letter

# --- CONFIGURAÇÃO VISUAL ---
st.set_page_config(page_title="Busca RMs | Deloitte", layout="wide")
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Open Sans', sans-serif; }
    [data-testid="stSidebar"] { background-color: #121212; }
    [data-testid="stSidebar"] * { color: #E0E0E0 !important; }
    h1, h2, h3 { color: #F0F0F0; }
    
    .resultado-card {
        background-color: #1E1E1E; padding: 15px; border-radius: 8px;
        border-left: 5px solid #86BC25; margin-bottom: 10px;
        display: flex; align-items: center; justify-content: space-between;
    }
    
    .color-dot {
        height: 20px; width: 20px; border-radius: 50%;
        border: 2px solid #FFF; display: inline-block;
        margin-right: 10px;
    }

    /* Estilo do Log de Leitura */
    .log-box { font-size: 0.8em; color: #888; margin-top: 10px; border-top: 1px solid #333; padding-top:10px; }
    .log-item { display: flex; justify-content: space-between; margin-bottom: 2px; }
    .log-warning { color: #FFCC00; }
    .log-ok { color: #86BC25; }

    div.stButton > button {
        background-color: #333333; color: #86BC25; border: 1px solid #86BC25; width: 100%;
    }
    div.stButton > button:hover { background-color: #86BC25; color: #000000; }
    </style>
""", unsafe_allow_html=True)

# --- CORES ---
CORES_VERDES = ["FF00FF00", "00FF00", "FF92D050", "92D050", "FFC6EFCE", "00B050", "FF00B050"]
CORES_VERMELHAS = ["FFFF0000", "FF0000", "FFFFC7CE", "FFC7CE", "C00000", "FFC00000"]

# --- FUNÇÃO DE CARREGAMENTO (AGORA UNIFICANDO CABEÇALHOS) ---
@st.cache_data(ttl=3600)
def carregar_planilha_inteligente(arquivo, usar_todas_abas, aba_especifica=None):
    xls = pd.ExcelFile(arquivo)
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    
    palavras_chave_header = ["RMT", "Código", "Revisão", "Status", "Descrição", "RM"]
    
    log_leitura = [] # Para guardar o relatório de quantas linhas leu de cada aba

    def processar_aba(nome_aba):
        ws = wb[nome_aba]
        linha_header = None
        colunas_validas_indices = None
        
        # 1. Tenta GPS do Filtro
        if ws.auto_filter.ref:
            try:
                min_c, min_r, max_c, max_r = range_boundaries(ws.auto_filter.ref)
                linha_header = min_r - 1
                use_cols = []
                colunas_validas_indices = []
                # Pega colunas visíveis dentro do filtro
                for col_idx in range(min_c, max_c + 1):
                    col_letter = get_column_letter(col_idx)
                    if not ws.column_dimensions[col_letter].hidden:
                        colunas_validas_indices.append(col_idx - 1)
                if not colunas_validas_indices: colunas_validas_indices = None
            except: pass

        # 2. Fallback: Busca por Palavra (Agora scanneia 50 linhas)
        if linha_header is None:
            df_preview = pd.read_excel(xls, sheet_name=nome_aba, header=None, nrows=50)
            for i, row in df_preview.iterrows():
                texto_linha = " ".join([str(val) for val in row.values]).lower()
                matches = sum(1 for palavra in palavras_chave_header if palavra.lower() in texto_linha)
                if matches >= 2:
                    linha_header = i
                    break
        
        # Se ainda não achou header, desiste dessa aba
        if linha_header is None:
            return None

        # 3. Ler Dados
        if colunas_validas_indices:
            df_final = pd.read_excel(xls, sheet_name=nome_aba, header=linha_header, usecols=colunas_validas_indices)
        else:
            df_final = pd.read_excel(xls, sheet_name=nome_aba, header=linha_header)

        # --- A MÁGICA DA UNIFICAÇÃO ---
        # Remove espaços extras dos nomes das colunas para "Status " virar "Status"
        df_final.columns = df_final.columns.astype(str).str.strip()
        
        # 4. Ler Cores (Usando busca por nome corrigido)
        col_status_idx = None
        col_prev_idx = None
        
        # Mapeia colunas pelo nome (agora limpo)
        for idx_df, col_name in enumerate(df_final.columns):
            c_name_lower = col_name.lower()
            if "status" in c_name_lower: 
                # Procura a coluna real no Excel
                for cell in ws[linha_header+1]:
                    if str(cell.value).strip() == col_name:
                        col_status_idx = cell.column; break
            if "previs" in c_name_lower or "entrega" in c_name_lower:
                for cell in ws[linha_header+1]:
                    if str(cell.value).strip() == col_name:
                        col_prev_idx = cell.column; break
        
        cores_bg = []
        cores_font = []
        
        for index in range(len(df_final)):
            excel_row = index + linha_header + 2
            try:
                # Fundo
                if col_status_idx:
                    c_bg = ws.cell(row=excel_row, column=col_status_idx).fill.start_color.index
                    cores_bg.append(str(c_bg) if c_bg and len(str(c_bg)) >= 6 else "NONE")
                else: cores_bg.append("NONE")

                # Fonte
                if col_prev_idx:
                    c_ft = ws.cell(row=excel_row, column=col_prev_idx).font.color
                    if c_ft and hasattr(c_ft, 'rgb') and c_ft.rgb: cores_font.append(str(c_ft.rgb))
                    else: cores_font.append("BLACK")
                else: cores_font.append("BLACK")
            except:
                cores_bg.append("ERROR"); cores_font.append("ERROR")

        df_final['[Cor Fundo]'] = cores_bg
        df_final['[Cor Fonte]'] = cores_font
        
        # Remove colunas vazias
        cols_validas = [c for c in df_final.columns if "Unnamed" not in c and c.lower() != "nan"]
        return df_final[cols_validas]

    # Lógica Central
    if usar_todas_abas:
        lista_dfs = []
        for aba in xls.sheet_names:
            try:
                # Ignora abas de sistema (ex: "Planilha1" vazia) se quiser
                df_temp = processar_aba(aba)
                if df_temp is not None and len(df_temp) > 0:
                    lista_dfs.append(df_temp)
                    log_leitura.append(f"{aba}: {len(df_temp)} linhas ✅")
                else:
                    log_leitura.append(f"{aba}: 0 linhas (Ignorada) ⚠️")
            except Exception as e:
                log_leitura.append(f"{aba}: Erro ❌")
        
        if not lista_dfs: return pd.DataFrame(), [], log_leitura
        
        # CONCATENAÇÃO INTELIGENTE (Junta colunas com nomes iguais)
        df_unificado = pd.concat(lista_dfs, ignore_index=True)
        return df_unificado, xls.sheet_names, log_leitura
    else:
        df = processar_aba(aba_especifica)
        log = [f"{aba_especifica}: {len(df) if df is not None else 0} linhas"]
        return df, xls.sheet_names, log

# --- INTERFACE ---
st.markdown("<h1><span style='color: #86BC25;'>Deloitte.</span> RMs <small>v7.0 (Unificador)</small></h1>", unsafe_allow_html=True)

st.sidebar.title("Controle")
if st.sidebar.button("🔄 Recarregar"):
    st.cache_data.clear()
    st.rerun()

# 1. SELETOR
arquivos = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]
arquivos.sort(key=lambda x: os.path.getmtime(x), reverse=True)

arquivo_selecionado = None
tab1, tab2 = st.sidebar.tabs(["Local", "Upload"])
with tab1:
    if arquivos:
        sel = st.selectbox("Arquivo:", arquivos)
        if st.checkbox("Carregar", value=True): arquivo_selecionado = sel
with tab2:
    up = st.file_uploader("Upload", type=["xlsx", "xls"])
    if up: arquivo_selecionado = up

# 2. PROCESSAMENTO
if arquivo_selecionado:
    try:
        xls_ref = pd.ExcelFile(arquivo_selecionado)
        abas = xls_ref.sheet_names
        df = None
        log_final = []

        st.sidebar.markdown("---")
        # Se tiver mais de uma aba, oferece opção de juntar
        if len(abas) > 1:
            modo = st.sidebar.radio("Leitura:", ["Juntar Todas (Recomendado)", "Escolher Uma"])
            
            if modo == "Juntar Todas (Recomendado)":
                with st.spinner("Unificando abas..."):
                    df, _, log_final = carregar_planilha_inteligente(arquivo_selecionado, True)
            else:
                aba = st.sidebar.selectbox("Qual aba?", abas)
                df, _, log_final = carregar_planilha_inteligente(arquivo_selecionado, False, aba)
        else:
            df, _, log_final = carregar_planilha_inteligente(arquivo_selecionado, False, abas[0])

        # --- EXIBIÇÃO DO LOG (NOVO) ---
        with st.sidebar.expander("📊 Relatório de Leitura", expanded=True):
            for linha in log_final:
                cor = "green" if "✅" in linha else "orange"
                st.markdown(f"<span style='color:{cor}; font-size:0.9em'>{linha}</span>", unsafe_allow_html=True)

        if df is None or df.empty:
            st.error("Não foi possível ler dados válidos. Verifique o relatório ao lado.")
            st.stop()

        colunas = df.columns.tolist()

        # --- BUSCA DE COLUNAS ---
        def achar(termos, evitar=None):
            if isinstance(termos, str): termos = [termos]
            for termo in termos:
                for i, col in enumerate(colunas):
                    c = str(col).lower()
                    if evitar and evitar.lower() in c: continue
                    if termo.lower() in c: return i
            return 0

        st.sidebar.markdown("### Mapeamento")
        col_rmt = st.sidebar.selectbox("RMT:", colunas, index=achar(["RMT", "Código", "Numero"]))
        col_rev = st.sidebar.selectbox("Revisão:", colunas, index=achar(["Rev", "R"], evitar="Previs"))
        col_st  = st.sidebar.selectbox("Status:", colunas, index=achar(["Status", "Situação"]))
        col_prev= st.sidebar.selectbox("Previsão:", colunas, index=achar(["Previs", "Entrega", "Prazo"]))
        
        st.divider()
        busca = st.text_input("Pesquisar RM:", placeholder="Código...").strip()

        if busca:
            mask = df[col_rmt].astype(str).str.strip().str.contains(busca, case=False, na=False)
            resultado = df[mask].copy()

            if not resultado.empty:
                try:
                    resultado['sort'] = resultado[col_rev].astype(str)
                    resultado = resultado.sort_values('sort')
                except: pass

                st.success(f"Encontrados: {len(resultado)}")
                texto_copia = ""
                data_hoje = datetime.now().strftime('%d/%m/%Y')
                rms = resultado[col_rmt].unique()

                for rm in rms:
                    texto_copia += f"{data_hoje}\n{rm}\n"
                    linhas = resultado[resultado[col_rmt] == rm]
                    
                    for _, row in linhas.iterrows():
                        rev = str(row[col_rev]).replace(".0", "") if pd.notna(row[col_rev]) else "-"
                        stat = str(row[col_st]) if pd.notna(row[col_st]) else "-"
                        
                        cor_bg = str(row.get('[Cor Fundo]', 'NONE'))
                        cor_ft = str(row.get('[Cor Fonte]', 'BLACK'))
                        
                        cor_html = "#333"
                        if len(cor_bg) == 8: cor_html = f"#{cor_bg[2:]}"
                        
                        raw_prev = row[col_prev]
                        if pd.notna(raw_prev):
                            v_str = str(raw_prev).strip()
                            ultima_data = v_str.split('\n')[-1].strip() if "\n" in v_str else v_str
                        else: ultima_data = None

                        indicador = ""
                        prev_final = ""

                        if cor_bg in CORES_VERDES:
                            prev_final = "ENTREGUE 🟢"
                            indicador = "✅ ENTREGUE"
                        elif cor_bg in CORES_VERMELHAS or cor_ft in CORES_VERMELHAS:
                            prev_final = f"{ultima_data} (ATRASADO 🔴)"
                            indicador = f"⚠️ {ultima_data}"
                        else:
                            if ultima_data:
                                try: prev_final = pd.to_datetime(ultima_data, dayfirst=True).strftime('%d/%m/%Y')
                                except: prev_final = ultima_data
                            else: prev_final = "Sem Previsão"
                            indicador = prev_final

                        st.markdown(f"""
                        <div class="resultado-card">
                            <div style="display:flex; align-items:center;">
                                <div class="color-dot" style="background-color: {cor_html};"></div>
                                <div><b style="color:#86BC25">{rm}</b> | REV: {rev} | {stat}</div>
                            </div>
                            <span style="color:#FFF; font-weight:bold;">{indicador}</span>
                        </div>
                        """, unsafe_allow_html=True)

                        texto_copia += f"REV {rev}\nStatus: {stat}\nPrevisão de entrega: {prev_final}\n\n"
                    
                    texto_copia += "--------------------\n"

                st.markdown("### Copiar (Ctrl+C)")
                st.text_area("Resultado:", value=texto_copia, height=300)

            else:
                st.warning("Nada encontrado.")
                if st.button("Gerar Linha Manual"):
                    st.code(f"{busca}\tDESCRIÇÃO\t0\tDiligenciamento\tVocê\tPendente\tFORNECEDOR\tDATA", language="text")

    except Exception as e:
        st.error(f"Erro Crítico: {e}")
else:
    st.info("Carregue uma planilha.")